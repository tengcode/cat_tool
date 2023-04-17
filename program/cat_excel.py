"""
Purpose    : cat excel sheet
Programmer : Bruce Ma
Start date : 2023-04-11
"""

import os
import openpyxl as op


def cat_multi(old_xlsx: op.Workbook,
              new_xlsx: op.Workbook,
              old_sheets: dict,
              new_sheets: dict
              ):
    # 合并到不同的sheet
    # 检查是否有同名sheet，若存在，计数后命名
    for old_sheet in old_sheets:
        sheet_count = new_sheets.get(old_sheet, 0)
        if sheet_count == 0:
            new_sheet = old_sheet
        else:
            new_sheet = f"{old_sheet}-{sheet_count}"
        new_sheets[old_sheet] = sheet_count + 1
        new_xlsx.create_sheet(new_sheet)
        old_ws = old_xlsx[old_sheet]
        new_ws = new_xlsx[new_sheet]
        for row in old_ws.rows:
            values = [cell.value for cell in row]
            new_ws.append(values)


def cat_single(old_xlsx: op.Workbook,
               new_xlsx: op.Workbook,
               old_sheets: dict,
               new_sheets: dict
               ):
    # 合并到同一个sheet
    for old_sheet in old_sheets:
        old_ws = old_xlsx[old_sheet]
        new_ws = new_xlsx.active
        for row in old_ws.rows:
            values = [cell.value for cell in row]
            new_ws.append(values)


def cat_excel(file_path: str,
              cat_type: str,
              new_file_name: str = '数据汇总',
              ):
    excel_file = []
    for file in os.listdir(file_path):
        if '.xlsx' in file:
            excel_file.append(file)

    new_file = f"{file_path}\\{new_file_name}.xlsx"
    new_xlsx = op.Workbook()
    new_sheets = {}
    for file in excel_file:
        old_file = f"{file_path}/{file}"
        old_xlsx = op.load_workbook(filename=old_file)
        old_sheets = old_xlsx.sheetnames
        if cat_type == 'multi':
            cat_multi(old_xlsx=old_xlsx,
                      new_xlsx=new_xlsx,
                      old_sheets=old_sheets,
                      new_sheets=new_sheets
                      )
        else:
            cat_single(old_xlsx=old_xlsx,
                       new_xlsx=new_xlsx,
                       old_sheets=old_sheets,
                       new_sheets=new_sheets
                       )
    new_xlsx.save(new_file)
